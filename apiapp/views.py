from django.shortcuts import render

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Trabajador, Expediente, Imagenes
from .serializers import TrabajadorSerializer, ExpedienteSerializer
import subprocess
import json
import os
import openpyxl
# Create your views here.
class ObtenerTrabajadorPorId(APIView):
    def post(self, request, id, tipo):
        try:
            expediente = Expediente.objects.get(id=id)
            imagenes = Imagenes.objects.filter(expediente=expediente.id)
            imagenes = [item.imagen for item in imagenes]
            tipo_lesion = [0,0,0,0]
            
            lesion = expediente.lesion.split("|")
            
            if(lesion[0].strip() == "Leve"):
                tipo_lesion[0] = 1
            if(lesion[0].strip() == "Grave"):
                tipo_lesion[1] = 1
            if(lesion[0].strip() == "MuyGrave"):
                tipo_lesion[2] = 1
            if(lesion[0].strip() == "Mortal"):
                tipo_lesion[3] = 1
            
            hora = ""
            fecha = ""
            
            if(expediente.fecha_suceso):
                hora = expediente.fecha_suceso.split("T")[1]
                fecha = expediente.fecha_suceso.split("T")[0]
            
            data = {
                "nombre_trabajador": f"{expediente.trabajador.apellido}, {expediente.trabajador.nombre}",
                "puesto_trabajo": expediente.puesto_trabajo,
                "edad": expediente.edad,
                "experiencia": expediente.trabajador.experiencia,
                "lugar_accidente": expediente.lugar_accidente,
                "hora_accidente": f"{hora}",
                "fecha_accidente": f"{fecha}",
                "lesion": lesion[1],
                "tipo_lesion": tipo_lesion,
                "lesionado_check": expediente.lesionado_check,
                "sexo": expediente.sexo,
                "descripcion_hechos": expediente.descripcion_hechos,
                "imagenes": str(imagenes),
                "id": expediente.id,
                "valoracionHechos": expediente.valoracion_hechos,
                "formas_accidente": expediente.formas_accidente,
                "analisis_causas": expediente.analisis_causas,
                "causas_accidente": expediente.causas_accidente,
                "aplicar_accion": expediente.aplicar_accion,
                "tipo_suceso": expediente.tipo_suceso,
                "creador": expediente.creador,
                "fecha_investigacion": expediente.fecha_investigacion
            }
            
            path_file = "json/expediente.json"
            
            with open(path_file, 'w') as f:
                json.dump(data, f, indent=4)
                
            
            if(tipo == "completo"):
                result2 = subprocess.run(['python', "-m", "pip", "install", "-r", "requirements.txt"])
                print(result2)
                result = subprocess.run(['python', os.path.abspath("generateWord.py")])
                print(result)
                return Response(status=status.HTTP_200_OK)
            
            if(tipo == "simplificado"):
                result2 = subprocess.run(['python', "-m", "pip", "install", "-r", "requirements.txt"])
                print(result2)
                result = subprocess.run(['python', os.path.abspath("generateWord2.py")])
                print(result)
                return Response(status=status.HTTP_200_OK)

            return Response({"msg": "El método indicado no existe"},status=status.HTTP_404_NOT_FOUND)
                    
        except Trabajador.DoesNotExist:
            return Response({'error': 'Expediente no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        

class ExportarExcel(APIView):
    def post(self, request):
        try:
            expediente = Expediente.objects.all()
            expedientes_catal = []
            expedientes_publindal = []
            for exp in expediente:
                if exp.empresa == "Catal":
                    expedientes_catal.append(exp)
                elif exp.empresa == "Publindal":
                    expedientes_publindal.append(exp)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            ws.title = "CATAL"
            ws['A1'] = "TIPO DE SUCESO"
            ws['B1'] = "EMPRESA"
            ws['C1'] = "FECHA"
            ws['D1'] = "HORA"
            ws['E1'] = "NOMBRE Y APELLIDOS"
            ws['F1'] = "PUESTO"
            ws['G1'] = "DESCIPCIÓN"
            ws['H1'] = "AGENTE"
            ws['I1'] = "PARTE DEL CUERPO"
            ws['J1'] = "FORMA DE PRODUCIRSE"
            
            
            ws2 = wb.create_sheet(title="PUBLINDAL")
            ws2['A1'] = "TIPO DE SUCESO"
            ws2['B1'] = "EMPRESA"
            ws2['C1'] = "FECHA"
            ws2['D1'] = "HORA"
            ws2['E1'] = "NOMBRE Y APELLIDOS"
            ws2['F1'] = "PUESTO"
            ws2['G1'] = "DESCIPCIÓN"
            ws2['H1'] = "AGENTE"
            ws2['I1'] = "PARTE DEL CUERPO"
            ws2['J1'] = "FORMA DE PRODUCIRSE"
            
            for index, exp in enumerate(expedientes_catal):
                index+=2
                cTipo = f"A{index}"
                cEmpresa = f"B{index}"
                cFecha = f"C{index}"
                cHora = f"D{index}"
                cNombreApellidos = f"E{index}"
                cPuesto = f"F{index}"
                cDescripcion = f"G{index}"
                cAgente = f"H{index}"
                cParteCuerpo = f"I{index}"
                cFormaProducirse = f"J{index}"
                
                ws[cTipo] = exp.tipo_suceso
                ws[cEmpresa] = exp.empresa
                ws[cFecha] = exp.fecha_suceso.split("T")[0]
                ws[cHora] = exp.fecha_suceso.split("T")[1]
                ws[cNombreApellidos] = f"{exp.trabajador.nombre} {exp.trabajador.apellido}"
                ws[cPuesto] = exp.puesto_trabajo
                ws[cDescripcion] = exp.lesion.split("|")[1]
                ws[cAgente] = exp.agente
                ws[cParteCuerpo] = exp.parte_cuerpo
                ws[cFormaProducirse] = exp.forma_producirse
                
            for index, exp in enumerate(expedientes_publindal):
                index+=2
                cTipo = f"A{index}"
                cEmpresa = f"B{index}"
                cFecha = f"C{index}"
                cHora = f"D{index}"
                cNombreApellidos = f"E{index}"
                cPuesto = f"F{index}"
                cDescripcion = f"G{index}"
                cAgente = f"H{index}"
                cParteCuerpo = f"I{index}"
                cFormaProducirse = f"J{index}"
                
                ws2[cTipo] = exp.tipo_suceso
                ws2[cEmpresa] = exp.empresa
                ws2[cFecha] = exp.fecha_suceso.split("T")[0]
                ws2[cHora] = exp.fecha_suceso.split("T")[1]
                ws2[cNombreApellidos] = f"{exp.trabajador.nombre} {exp.trabajador.apellido}"
                ws2[cPuesto] = exp.puesto_trabajo
                ws2[cDescripcion] = exp.lesion.split("|")[1]
                ws2[cAgente] = exp.agente
                ws2[cParteCuerpo] = exp.parte_cuerpo
                ws2[cFormaProducirse] = exp.forma_producirse
                
            
            doc = os.path.abspath('media/excel/expedientes.xlsx')
            wb.save(doc)
            
            return Response(status=status.HTTP_200_OK)
        except Exception as e: 
            print(e)
            return Response({'error': 'Algo ha salido mal'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
        